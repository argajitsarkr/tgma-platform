"""ID Allocation — pre-allocate tracking IDs for field workers before they go to field."""

import io
import base64
from datetime import date
from flask import Blueprint, render_template, request, redirect, url_for, flash, Response
from flask_login import login_required, current_user

from app.extensions import db
from app.models import IdAllocation, Participant
from app.utils.helpers import generate_tracking_id
from app.utils.decorators import role_required

ids_bp = Blueprint('ids', __name__, url_prefix='/ids')


def _get_next_seq(district, gender):
    """Return the next available sequence number for a district/gender combination."""
    max_alloc = db.session.query(db.func.max(IdAllocation.tracking_id)).filter(
        IdAllocation.tracking_id.like(f'TGMA-{district}-{gender}-%')
    ).scalar()
    max_part = db.session.query(db.func.max(Participant.tracking_id)).filter(
        Participant.tracking_id.like(f'TGMA-{district}-{gender}-%')
    ).scalar()

    max_seq = 0
    for tid in [max_alloc, max_part]:
        if tid:
            try:
                seq = int(tid.split('-')[-1])
                max_seq = max(max_seq, seq)
            except (ValueError, IndexError):
                pass
    return max_seq + 1


@ids_bp.route('/')
@login_required
def index():
    """Show allocation dashboard — current sequence status per district/gender."""
    summaries = {}
    for district in ['WT', 'ST', 'DL']:
        for gender in ['M', 'F']:
            key = f"{district}-{gender}"
            next_seq = _get_next_seq(district, gender)
            summaries[key] = {
                'district': district,
                'gender': gender,
                'last_sequence': next_seq - 1,
                'next_sequence': next_seq,
            }

    # Recent allocations
    recent = IdAllocation.query.order_by(IdAllocation.created_at.desc()).limit(50).all()

    # Group consecutive allocations into batches by (worker, date, district, gender)
    batches = []
    for a in recent:
        parts = a.tracking_id.split('-')
        if len(parts) != 4:
            continue
        d, g = parts[1], parts[2]
        key = (a.allocated_to, a.allocated_date, d, g)
        if batches and batches[-1]['key'] == key:
            batches[-1]['ids'].append(a.tracking_id)
        else:
            batches.append({
                'key': key,
                'worker': a.allocated_to,
                'date': a.allocated_date,
                'district': d,
                'gender': g,
                'ids': [a.tracking_id],
            })
    for b in batches:
        b['ids'].sort()
        b['start'] = b['ids'][0]
        b['end'] = b['ids'][-1]
        b['count'] = len(b['ids'])

    return render_template('ids/allocate.html', summaries=summaries,
                           recent=recent, batches=batches)


@ids_bp.route('/allocate', methods=['POST'])
@login_required
@role_required('pi', 'co_pi', 'bioinformatician')
def allocate():
    """Bulk-allocate a batch of IDs for a field worker."""
    district = request.form.get('district', '').upper()
    gender = request.form.get('gender', '').upper()
    field_worker = request.form.get('field_worker', '').strip()
    try:
        count = max(1, min(50, int(request.form.get('count', 1))))
    except (ValueError, TypeError):
        count = 1

    if district not in ('WT', 'ST', 'DL'):
        flash('Invalid district.', 'danger')
        return redirect(url_for('ids.index'))
    if gender not in ('M', 'F'):
        flash('Invalid gender.', 'danger')
        return redirect(url_for('ids.index'))

    start_seq = _get_next_seq(district, gender)
    allocated_ids = []

    for i in range(count):
        seq = start_seq + i
        tracking_id = generate_tracking_id(district, gender, seq)
        alloc = IdAllocation(
            tracking_id=tracking_id,
            allocated_date=date.today(),
            allocated_to=field_worker or None,
            status='allocated',
        )
        db.session.add(alloc)
        allocated_ids.append(tracking_id)

    db.session.commit()

    flash(f'Allocated {count} ID{"s" if count > 1 else ""}: '
          f'{allocated_ids[0]} → {allocated_ids[-1]}', 'success')

    # Redirect to print-ready label sheet if batch > 1
    if count > 1:
        return redirect(url_for('ids.print_labels',
                                start=allocated_ids[0], end=allocated_ids[-1],
                                worker=field_worker))
    return redirect(url_for('ids.index'))


@ids_bp.route('/labels')
@login_required
def print_labels():
    """Print-ready label sheet for a batch of allocated IDs."""
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    worker = request.args.get('worker', '')

    # Collect all IDs in the range from allocation table
    ids = IdAllocation.query.filter(
        IdAllocation.tracking_id >= start,
        IdAllocation.tracking_id <= end,
        IdAllocation.allocated_to == (worker or None)
            if worker else IdAllocation.tracking_id >= start,
    ).order_by(IdAllocation.tracking_id).all()

    # Fallback: generate list from start/end directly
    if not ids and start and end:
        try:
            parts_s = start.split('-')
            parts_e = end.split('-')
            district, gender = parts_s[1], parts_s[2]
            seq_s = int(parts_s[3])
            seq_e = int(parts_e[3])
            label_ids = [generate_tracking_id(district, gender, s)
                         for s in range(seq_s, seq_e + 1)]
        except (IndexError, ValueError):
            label_ids = []
    else:
        label_ids = [a.tracking_id for a in ids]

    return render_template('ids/labels.html',
                           label_ids=label_ids,
                           worker=worker,
                           today=date.today())


# -- Label Kit definition: 10 labels per participant --
# Blood vials removed — diagnostics company requires handwritten name/age/sex.
LABEL_KIT = [
    # (suffix, description, category)
    ('',      'Participant Folder',    'Folder'),
    ('-STL',  'Fecal Sample',          'Fecal'),
    ('-SLV1', 'Morning (6-8 AM)',      'Saliva'),
    ('-SLV2', 'Noon (12-1 PM)',        'Saliva'),
    ('-SLV3', 'Evening (5-6 PM)',      'Saliva'),
    ('-SLV4', 'Night (10-11 PM)',      'Saliva'),
    ('-DOC',  'Consent Form',          'Document'),
    ('-DOC',  'Assent Form',           'Document'),
    ('-DOC',  'Information Sheet',     'Document'),
    ('-DOC',  'Questionnaire',         'Document'),
]


@ids_bp.route('/label-kit/excel')
@login_required
def label_kit_excel():
    """Generate Excel file with all label data for the Seznik thermal printer app.

    Query params: start, end (tracking ID range), worker (optional).
    Output: Excel file with columns: Barcode, Label_Text, Description, Category, Participant.
    The field worker imports this into the Seznik app via 'Print Excel'.
    """
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    worker = request.args.get('worker', '')

    # Resolve the list of tracking IDs (same logic as print_labels)
    ids = IdAllocation.query.filter(
        IdAllocation.tracking_id >= start,
        IdAllocation.tracking_id <= end,
    ).order_by(IdAllocation.tracking_id).all()

    if not ids and start and end:
        try:
            parts_s = start.split('-')
            parts_e = end.split('-')
            district, gender = parts_s[1], parts_s[2]
            seq_s = int(parts_s[3])
            seq_e = int(parts_e[3])
            label_ids = [generate_tracking_id(district, gender, s)
                         for s in range(seq_s, seq_e + 1)]
        except (IndexError, ValueError):
            label_ids = []
    else:
        label_ids = [a.tracking_id for a in ids]

    if not label_ids:
        flash('No IDs found for the given range.', 'danger')
        return redirect(url_for('ids.index'))

    # Build Excel workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = 'Label Kit'

    # Header styling
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Category fills for visual grouping
    cat_fills = {
        'Folder':   PatternFill(start_color='D8F3DC', end_color='D8F3DC', fill_type='solid'),
        'Blood':    PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
        'Fecal':    PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid'),
        'Saliva':   PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'),
        'Document': PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid'),
    }

    # Write headers
    headers = ['Barcode', 'Label_Text', 'Description', 'Category', 'Participant_ID',
               'Field_Worker', 'Date']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write label rows
    row_num = 2
    for tracking_id in label_ids:
        for suffix, description, category in LABEL_KIT:
            barcode_value = f'{tracking_id}{suffix}'
            if category == 'Folder':
                label_text = tracking_id
                description = 'Participant Folder'
            else:
                label_text = f'{tracking_id}{suffix}'

            ws.cell(row=row_num, column=1, value=barcode_value).border = thin_border
            ws.cell(row=row_num, column=2, value=label_text).border = thin_border
            ws.cell(row=row_num, column=3, value=description).border = thin_border
            ws.cell(row=row_num, column=4, value=category).border = thin_border
            ws.cell(row=row_num, column=5, value=tracking_id).border = thin_border
            ws.cell(row=row_num, column=6, value=worker or '—').border = thin_border
            ws.cell(row=row_num, column=7, value=date.today().isoformat()).border = thin_border

            # Apply category color
            fill = cat_fills.get(category)
            if fill:
                for col in range(1, 8):
                    ws.cell(row=row_num, column=col).fill = fill

            row_num += 1

    # Column widths
    widths = [28, 28, 22, 12, 22, 16, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f'A1:G{row_num - 1}'

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'TGMA_Label_Kit_{start}_to_{end}_{date.today().isoformat()}.xlsx'

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@ids_bp.route('/thermal-labels/<tracking_id>')
@login_required
def thermal_labels(tracking_id):
    """Print-ready page for 10 thermal labels (50x30mm page, 25mm sticker) with QR codes.

    Each QR encodes a URL to the participant's detail page on the platform.
    Scanning a label with any phone camera opens the participant record directly.
    """
    import qrcode
    from qrcode.constants import ERROR_CORRECT_M

    # Build base URL for participant detail links
    base_url = request.host_url.rstrip('/')

    labels = []
    for suffix, description, category in LABEL_KIT:
        barcode_value = f'{tracking_id}{suffix}'

        # QR encodes a direct URL to the participant's page
        qr_url = f'{base_url}/participants/{tracking_id}'

        qr = qrcode.QRCode(
            version=None,
            error_correction=ERROR_CORRECT_M,
            box_size=8,
            border=2,
        )
        qr.add_data(qr_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        b64 = base64.b64encode(buf.getvalue()).decode('ascii')

        labels.append({
            'barcode_value': barcode_value,
            'qr_b64': b64,
            'description': description,
            'category': category,
            'tracking_id': tracking_id,
        })

    return render_template('ids/thermal_labels.html',
                           labels=labels,
                           tracking_id=tracking_id)


@ids_bp.route('/delete-allocation/<tracking_id>', methods=['POST'])
@login_required
@role_required('pi', 'co_pi', 'bioinformatician')
def delete_allocation(tracking_id):
    """Delete a single ID allocation — only if no Participant uses it."""
    alloc = db.session.get(IdAllocation, tracking_id)
    if not alloc:
        flash('Allocation not found.', 'danger')
        return redirect(url_for('ids.index'))

    if db.session.get(Participant, tracking_id):
        flash(f'Cannot delete {tracking_id} — a participant record exists with this ID.', 'warning')
        return redirect(url_for('ids.index'))

    db.session.delete(alloc)
    db.session.commit()
    flash(f'Allocation {tracking_id} deleted.', 'success')
    return redirect(url_for('ids.index'))


@ids_bp.route('/delete-batch', methods=['POST'])
@login_required
@role_required('pi', 'co_pi', 'bioinformatician')
def delete_batch():
    """Delete all ID allocations in a batch — refuses if any have a Participant."""
    ids = request.form.getlist('tracking_ids')
    if not ids:
        flash('No IDs specified.', 'danger')
        return redirect(url_for('ids.index'))

    blocked = []
    to_delete = []
    for tid in ids:
        alloc = db.session.get(IdAllocation, tid)
        if not alloc:
            continue
        if db.session.get(Participant, tid):
            blocked.append(tid)
        else:
            to_delete.append(alloc)

    if blocked:
        flash(f'Cannot delete {len(blocked)} ID(s) with existing participants: {", ".join(blocked)}', 'warning')
        return redirect(url_for('ids.index'))

    for alloc in to_delete:
        db.session.delete(alloc)
    db.session.commit()
    flash(f'Deleted {len(to_delete)} ID allocation(s).', 'success')
    return redirect(url_for('ids.index'))
